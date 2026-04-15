from __future__ import annotations

import argparse
import os
import time
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import requests
import tushare as ts
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


API_URL = "http://api.waditu.com"
FLOAT_MV_THRESHOLD_WAN = 350_000
GAIN_THRESHOLD = 0.05
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial", size=10)


@dataclass(frozen=True)
class Interval:
    index: int
    start_date: str
    end_date: str

    @property
    def label(self) -> str:
        return f"I{self.index:02d} {self.start_date}~{self.end_date}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="筛选 OAMV 死叉区间内强势股票")
    parser.add_argument(
        "--signal-file",
        type=Path,
        default=Path("analysis/OAMV_MACD.xlsx"),
        help="OAMV_MACD 信号 Excel 路径",
    )
    parser.add_argument(
        "--start-date",
        default="2024-01-01",
        help="分析开始日期，格式 YYYY-MM-DD",
    )
    parser.add_argument(
        "--end-date",
        default="2026-03-20",
        help="分析结束日期，格式 YYYY-MM-DD",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("analysis/OAMV_deadcross_screen_20240101_20260320.xlsx"),
        help="输出 Excel 路径",
    )
    parser.add_argument(
        "--max-intervals",
        type=int,
        default=0,
        help="仅调试用，限制处理前 N 个死叉区间；0 表示全部",
    )
    return parser.parse_args()


def get_token() -> str:
    token = os.getenv("TUSHARE_TOKEN") or ts.get_token()
    if not token:
        raise RuntimeError("未找到 TUSHARE_TOKEN，也没有本地 tushare token 配置。")
    return token


def create_session() -> requests.Session:
    session = requests.Session()
    session.trust_env = False
    session.headers.update(
        {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
        }
    )
    return session


def tushare_call(
    session: requests.Session,
    token: str,
    api_name: str,
    params: dict[str, object],
    fields: str,
    retries: int = 4,
) -> pd.DataFrame:
    payload = {
        "api_name": api_name,
        "token": token,
        "params": params,
        "fields": fields,
    }
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            response = session.post(API_URL, json=payload, timeout=45)
            response.raise_for_status()
            response.encoding = "utf-8"
            data = response.json()
            if data.get("code") != 0:
                raise RuntimeError(f"{api_name} 返回错误: {data.get('msg')}")
            data_block = data.get("data") or {}
            rows = data_block.get("items") or []
            columns = data_block.get("fields") or []
            return pd.DataFrame(rows, columns=columns)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(1.0 + attempt * 1.5)
    raise RuntimeError(f"Tushare 接口 {api_name} 调用失败: {last_error}")


def read_signal_events(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet1")
    expected = ["日期", "活跃市值状态"]
    missing = set(expected) - set(df.columns)
    if missing:
        raise ValueError(f"信号表缺少必要列: {', '.join(sorted(missing))}")
    df = df[expected].copy()
    df["日期"] = pd.to_datetime(df["日期"])
    df = df.dropna(subset=["日期", "活跃市值状态"]).sort_values("日期")
    return df


def build_intervals(events: pd.DataFrame, start_date: str, end_date: str) -> list[Interval]:
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    current_death: pd.Timestamp | None = None
    intervals: list[Interval] = []

    for _, row in events.iterrows():
        trade_day = row["日期"]
        status = str(row["活跃市值状态"]).strip()
        if trade_day < start_ts or trade_day > end_ts:
            continue
        if status == "死叉":
            current_death = trade_day
            continue
        if status == "金叉" and current_death is not None and trade_day >= current_death:
            intervals.append(
                Interval(
                    index=len(intervals) + 1,
                    start_date=current_death.strftime("%Y%m%d"),
                    end_date=trade_day.strftime("%Y%m%d"),
                )
            )
            current_death = None

    return intervals


def fetch_trade_calendar(
    session: requests.Session,
    token: str,
    start_date: str,
    end_date: str,
) -> list[str]:
    cal = tushare_call(
        session,
        token,
        api_name="trade_cal",
        params={"exchange": "SSE", "start_date": start_date, "end_date": end_date, "is_open": "1"},
        fields="cal_date,is_open",
    )
    if cal.empty:
        raise RuntimeError("交易日历为空，无法继续分析。")
    return cal["cal_date"].astype(str).sort_values().tolist()


def fetch_stock_master(session: requests.Session, token: str) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for status in ("L", "D", "P"):
        df = tushare_call(
            session,
            token,
            api_name="stock_basic",
            params={"exchange": "", "list_status": status},
            fields="ts_code,symbol,name,industry,market,list_date,delist_date",
        )
        frames.append(df)
    stock = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["ts_code"], keep="first")
    return stock


def interval_trade_dates(all_trade_dates: list[str], interval: Interval) -> list[str]:
    return [d for d in all_trade_dates if interval.start_date <= d <= interval.end_date]


def fetch_daily_for_dates(
    session: requests.Session,
    token: str,
    trade_dates: list[str],
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    total = len(trade_dates)
    for idx, trade_date in enumerate(trade_dates, start=1):
        if idx == 1 or idx == total or idx % 20 == 0:
            print(f"[daily] {idx}/{total} {trade_date}", flush=True)
        df = tushare_call(
            session,
            token,
            api_name="daily",
            params={"trade_date": trade_date},
            fields="ts_code,trade_date,close",
        )
        if not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame(columns=["ts_code", "trade_date", "close"])
    return pd.concat(frames, ignore_index=True)


def fetch_adj_for_dates(
    session: requests.Session,
    token: str,
    trade_dates: list[str],
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    total = len(trade_dates)
    for idx, trade_date in enumerate(trade_dates, start=1):
        if idx == 1 or idx == total or idx % 20 == 0:
            print(f"[adj]   {idx}/{total} {trade_date}", flush=True)
        df = tushare_call(
            session,
            token,
            api_name="adj_factor",
            params={"trade_date": trade_date},
            fields="ts_code,trade_date,adj_factor",
        )
        if not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame(columns=["ts_code", "trade_date", "adj_factor"])
    return pd.concat(frames, ignore_index=True)


def fetch_start_caps(
    session: requests.Session,
    token: str,
    trade_date: str,
) -> pd.DataFrame:
    df = tushare_call(
        session,
        token,
        api_name="daily_basic",
        params={"trade_date": trade_date},
        fields="ts_code,trade_date,circ_mv",
    )
    if df.empty:
        return df
    df["circ_mv"] = pd.to_numeric(df["circ_mv"], errors="coerce")
    df = df[df["circ_mv"] > FLOAT_MV_THRESHOLD_WAN].copy()
    return df


def summarize_interval(
    interval: Interval,
    start_caps: pd.DataFrame,
    daily_df: pd.DataFrame,
    adj_df: pd.DataFrame,
    stock_master: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, object]]:
    if daily_df.empty or adj_df.empty or start_caps.empty:
        return pd.DataFrame(), {"区间": interval.label, "候选股票数": 0, "涨幅>5%股票数": 0}

    merged = daily_df.merge(adj_df, on=["ts_code", "trade_date"], how="inner")
    merged["close"] = pd.to_numeric(merged["close"], errors="coerce")
    merged["adj_factor"] = pd.to_numeric(merged["adj_factor"], errors="coerce")
    merged = merged.dropna(subset=["close", "adj_factor"]).copy()
    merged["adj_close"] = merged["close"] * merged["adj_factor"]
    merged = merged.sort_values(["ts_code", "trade_date"])
    merged = merged[merged["ts_code"].isin(start_caps["ts_code"])]

    if merged.empty:
        return pd.DataFrame(), {"区间": interval.label, "候选股票数": 0, "涨幅>5%股票数": 0}

    first_rows = merged.groupby("ts_code", as_index=False).first()
    last_rows = merged.groupby("ts_code", as_index=False).last()

    interval_df = first_rows[["ts_code", "trade_date", "adj_close"]].rename(
        columns={"trade_date": "起始交易日", "adj_close": "起始复权价"}
    )
    interval_df = interval_df.merge(
        last_rows[["ts_code", "trade_date", "adj_close"]].rename(
            columns={"trade_date": "结束交易日", "adj_close": "结束复权价"}
        ),
        on="ts_code",
        how="inner",
    )
    interval_df = interval_df.merge(
        start_caps[["ts_code", "circ_mv"]].rename(columns={"circ_mv": "死叉首日流通市值(万元)"}),
        on="ts_code",
        how="inner",
    )
    interval_df["区间涨幅"] = interval_df["结束复权价"] / interval_df["起始复权价"] - 1
    interval_df = interval_df[interval_df["区间涨幅"] > GAIN_THRESHOLD].copy()
    if interval_df.empty:
        return pd.DataFrame(), {"区间": interval.label, "候选股票数": int(len(start_caps)), "涨幅>5%股票数": 0}

    interval_df = interval_df.merge(stock_master, on="ts_code", how="left")
    interval_df["区间"] = interval.label
    interval_df["死叉日"] = pd.to_datetime(interval.start_date).strftime("%Y-%m-%d")
    interval_df["金叉日"] = pd.to_datetime(interval.end_date).strftime("%Y-%m-%d")
    interval_df["死叉首日流通市值(亿元)"] = interval_df["死叉首日流通市值(万元)"] / 10_000
    interval_df["区间涨幅"] = interval_df["区间涨幅"].astype(float)

    summary = {
        "区间": interval.label,
        "死叉日": pd.to_datetime(interval.start_date).strftime("%Y-%m-%d"),
        "金叉日": pd.to_datetime(interval.end_date).strftime("%Y-%m-%d"),
        "候选股票数": int(len(start_caps)),
        "涨幅>5%股票数": int(len(interval_df)),
        "区间内平均涨幅": float(interval_df["区间涨幅"].mean()),
    }
    keep_cols = [
        "区间",
        "死叉日",
        "金叉日",
        "ts_code",
        "symbol",
        "name",
        "industry",
        "market",
        "死叉首日流通市值(亿元)",
        "起始交易日",
        "结束交易日",
        "起始复权价",
        "结束复权价",
        "区间涨幅",
    ]
    return interval_df[keep_cols].sort_values("区间涨幅", ascending=False), summary


def score_from_count(count: int) -> int:
    if count > 8:
        return 2
    if count > 4:
        return 1
    return 0


def score_from_mean_gain(mean_gain: float) -> int:
    if mean_gain > 0.15:
        return 2
    if mean_gain > 0.10:
        return 1
    return 0


def build_summary_sheet(interval_hits: pd.DataFrame) -> pd.DataFrame:
    if interval_hits.empty:
        return pd.DataFrame(
            columns=[
                "排名",
                "总分",
                "次数得分",
                "平均涨幅得分",
                "上涨区间次数",
                "平均区间涨幅",
                "ts_code",
                "symbol",
                "name",
                "industry",
                "market",
                "命中区间",
            ]
        )

    summary = (
        interval_hits.groupby(["ts_code", "symbol", "name", "industry", "market"], dropna=False)
        .agg(
            上涨区间次数=("区间", "count"),
            平均区间涨幅=("区间涨幅", "mean"),
            最大区间涨幅=("区间涨幅", "max"),
            命中区间=("区间", lambda s: " | ".join(s)),
        )
        .reset_index()
    )
    summary = summary[summary["上涨区间次数"] > 3].copy()
    summary["次数得分"] = summary["上涨区间次数"].apply(score_from_count)
    summary["平均涨幅得分"] = summary["平均区间涨幅"].apply(score_from_mean_gain)
    summary["总分"] = summary["次数得分"] + summary["平均涨幅得分"]
    summary = summary.sort_values(
        ["总分", "上涨区间次数", "平均区间涨幅", "最大区间涨幅"],
        ascending=[False, False, False, False],
    ).reset_index(drop=True)
    summary.insert(0, "排名", summary.index + 1)
    return summary[
        [
            "排名",
            "总分",
            "次数得分",
            "平均涨幅得分",
            "上涨区间次数",
            "平均区间涨幅",
            "最大区间涨幅",
            "ts_code",
            "symbol",
            "name",
            "industry",
            "market",
            "命中区间",
        ]
    ]


def autosize_and_style(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
                if cell.column_letter in {"F", "G", "H", "I", "J", "K", "L", "M"} and isinstance(cell.value, (int, float)):
                    pass
        for column in ws.columns:
            letter = column[0].column_letter
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)
    wb.save(workbook_path)


def main() -> int:
    args = parse_args()
    signal_path = args.signal_file
    output_path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    events = read_signal_events(signal_path)
    intervals = build_intervals(events, args.start_date, args.end_date)
    if args.max_intervals > 0:
        intervals = intervals[: args.max_intervals]
    if not intervals:
        raise RuntimeError("在指定日期范围内没有找到完整的死叉到金叉区间。")

    token = get_token()
    session = create_session()
    stock_master = fetch_stock_master(session, token)
    trade_dates = fetch_trade_calendar(
        session,
        token,
        start_date=intervals[0].start_date,
        end_date=intervals[-1].end_date,
    )

    interval_frames: list[pd.DataFrame] = []
    interval_summaries: list[dict[str, object]] = []

    for interval in intervals:
        dates = interval_trade_dates(trade_dates, interval)
        print(f"处理区间 {interval.label}，交易日 {len(dates)} 天", flush=True)
        start_caps = fetch_start_caps(session, token, interval.start_date)
        daily_df = fetch_daily_for_dates(session, token, dates)
        adj_df = fetch_adj_for_dates(session, token, dates)
        interval_df, interval_summary = summarize_interval(interval, start_caps, daily_df, adj_df, stock_master)
        interval_frames.append(interval_df)
        interval_summaries.append(interval_summary)

    interval_hits = pd.concat(interval_frames, ignore_index=True) if interval_frames else pd.DataFrame()
    summary_sheet = build_summary_sheet(interval_hits)
    interval_summary_sheet = pd.DataFrame(interval_summaries)

    if not interval_hits.empty:
        interval_hits["区间涨幅"] = interval_hits["区间涨幅"].map(lambda x: f"{x:.2%}")
    if not summary_sheet.empty:
        summary_sheet["平均区间涨幅"] = summary_sheet["平均区间涨幅"].map(lambda x: f"{x:.2%}")
        summary_sheet["最大区间涨幅"] = summary_sheet["最大区间涨幅"].map(lambda x: f"{x:.2%}")
    if not interval_summary_sheet.empty and "区间内平均涨幅" in interval_summary_sheet.columns:
        interval_summary_sheet["区间内平均涨幅"] = interval_summary_sheet["区间内平均涨幅"].map(
            lambda x: f"{x:.2%}" if pd.notna(x) else ""
        )

    rule_sheet = pd.DataFrame(
        {
            "规则": [
                "流通市值门槛",
                "有效上涨区间定义",
                "最终入选门槛",
                "次数得分",
                "平均涨幅得分",
                "总分",
            ],
            "说明": [
                "死叉首日流通市值 > 35亿元",
                "死叉到下一次金叉区间内，按个股区间首个可交易日到末个可交易日计算复权涨幅，且涨幅 > 5%",
                "上涨区间次数 > 3 次",
                ">4次记1分，>8次记2分",
                "平均区间涨幅 >10%记1分，>15%记2分",
                "总分 = 次数得分 + 平均涨幅得分，并按总分/次数/平均涨幅降序排序",
            ],
        }
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_sheet.to_excel(writer, sheet_name="汇总排名", index=False)
        interval_hits.to_excel(writer, sheet_name="区间明细", index=False)
        interval_summary_sheet.to_excel(writer, sheet_name="区间统计", index=False)
        rule_sheet.to_excel(writer, sheet_name="评分规则", index=False)

    autosize_and_style(output_path)
    print(f"已输出: {output_path.resolve()}", flush=True)
    print(f"完整死叉区间数: {len(intervals)}", flush=True)
    print(f"最终入选股票数: {len(summary_sheet)}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
